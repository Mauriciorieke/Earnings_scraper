"""Excel output module for 3-statement financial models.

Generates a formatted Excel workbook with separate tabs for:
- Income Statement (all discovered line items)
- Balance Sheet (all discovered line items)
- Cash Flow Statement (all discovered line items)
- Summary / Dashboard

Handles any number of line items — auto-detects totals and subtotals
from the label text rather than relying on a hardcoded list.
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, ".")
from config import OUTPUT_DIR


# ------------------------------------------------------------------
# Styling constants
# ------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="808080"),
)
NUMBER_FORMAT = '#,##0'
EPS_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'

# Regex patterns for auto-detecting total / subtotal rows
_TOTAL_PATTERN = re.compile(
    r"(?i)^(?:total|net income|net loss|gross profit|operating income|"
    r"stockholders.?equity|liabilities and|comprehensive income|"
    r"cash.*(?:operations|investing|financing)|"
    r"net (?:cash|change))",
)

# Regex patterns for per-share items
_PER_SHARE_PATTERN = re.compile(
    r"(?i)(?:earnings per share|per share|EPS|\[USD/shares\])",
)

# Regex patterns for share count items
_SHARES_PATTERN = re.compile(
    r"(?i)(?:\[shares\]|shares outstanding|weighted average.*shares)",
)


def _is_total_row(label):
    """Auto-detect whether a line item label is a total/subtotal."""
    return bool(_TOTAL_PATTERN.search(label))


def _get_number_format(label):
    """Pick the right number format based on the line item label."""
    if _PER_SHARE_PATTERN.search(label):
        return EPS_FORMAT
    if _SHARES_PATTERN.search(label):
        return '#,##0'
    return NUMBER_FORMAT


def _style_sheet(ws, df, sheet_title):
    """Apply formatting to a worksheet containing a financial statement."""
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(df.columns) + 1, 30))
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="left")

    # Header row (period dates)
    header_row = 3
    ws.cell(row=header_row, column=1, value="Line Item").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left")

    for col_idx, period in enumerate(df.columns, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=str(period))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (line_item, row_data) in enumerate(df.iterrows(), start=header_row + 1):
        label_cell = ws.cell(row=row_idx, column=1, value=line_item)

        is_total = _is_total_row(line_item)
        num_fmt = _get_number_format(line_item)

        if is_total:
            label_cell.font = TOTAL_FONT
            label_cell.fill = TOTAL_FILL
            label_cell.border = THIN_BORDER
        else:
            label_cell.font = DATA_FONT

        for col_idx, period in enumerate(df.columns, start=2):
            val = row_data[period]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.notna(val):
                cell.value = val
                cell.number_format = num_fmt
            else:
                cell.value = ""

            cell.alignment = Alignment(horizontal="right")
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
            else:
                cell.font = DATA_FONT

    # Column widths
    # Auto-size column A based on longest label
    max_label_len = max((len(str(idx)) for idx in df.index), default=20)
    ws.column_dimensions["A"].width = min(max_label_len + 4, 60)
    for col_idx in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze panes: freeze the line item column and header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)


def _add_summary_sheet(wb, statements, company_name):
    """Add a summary/dashboard sheet with key metrics pulled dynamically."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=f"{company_name} — Financial Summary")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="2F5496")

    ws.cell(row=3, column=1, value="Key Metrics (Most Recent Annual Periods)")
    ws.cell(row=3, column=1).font = Font(name="Calibri", bold=True, size=12)

    # Key metric label patterns to pull from each statement
    key_metric_patterns = {
        "Income Statement": [
            r"(?i)^revenue",
            r"(?i)^gross profit",
            r"(?i)^operating income",
            r"(?i)^net income",
            r"(?i)earnings per share.*diluted",
            r"(?i)^research and development",
            r"(?i)^selling.*general.*admin",
        ],
        "Balance Sheet": [
            r"(?i)^cash",
            r"(?i)^total.*current.*assets",
            r"(?i)^assets$",
            r"(?i)^total.*current.*liabilities",
            r"(?i)^liabilities$",
            r"(?i)^stockholders.*equity$",
            r"(?i)^long.term.*debt",
        ],
        "Cash Flow Statement": [
            r"(?i)net cash.*operating",
            r"(?i)payments.*acquire.*property",
            r"(?i)net cash.*investing",
            r"(?i)net cash.*financing",
            r"(?i)payments.*dividends",
            r"(?i)payments.*repurchase.*common",
        ],
    }

    current_row = 5
    for statement_name, patterns in key_metric_patterns.items():
        ws.cell(row=current_row, column=1, value=statement_name)
        ws.cell(row=current_row, column=1).font = SUBHEADER_FONT
        ws.cell(row=current_row, column=1).fill = SUBHEADER_FILL
        for c in range(2, 8):
            ws.cell(row=current_row, column=c).fill = SUBHEADER_FILL

        df = statements.get(statement_name)
        if df is None or df.empty:
            current_row += 1
            ws.cell(row=current_row, column=1, value="(no data)")
            current_row += 2
            continue

        # Use last 5 periods
        recent_periods = list(df.columns[-5:])

        # Write period headers
        for col_idx, period in enumerate(recent_periods, start=2):
            ws.cell(row=current_row, column=col_idx, value=str(period))
            ws.cell(row=current_row, column=col_idx).font = SUBHEADER_FONT
            ws.cell(row=current_row, column=col_idx).fill = SUBHEADER_FILL
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center")

        current_row += 1

        # Find matching line items for each pattern
        for pattern in patterns:
            regex = re.compile(pattern)
            matched = [idx for idx in df.index if regex.search(idx)]
            if not matched:
                continue
            metric = matched[0]  # Take first match
            ws.cell(row=current_row, column=1, value=metric).font = DATA_FONT
            num_fmt = _get_number_format(metric)
            for col_idx, period in enumerate(recent_periods, start=2):
                val = df.at[metric, period]
                cell = ws.cell(row=current_row, column=col_idx)
                if pd.notna(val):
                    cell.value = val
                    cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
                cell.font = DATA_FONT
            current_row += 1

        current_row += 1  # blank row between sections

    # Forecast placeholder section
    current_row += 1
    ws.cell(row=current_row, column=1, value="FORECAST ASSUMPTIONS")
    ws.cell(row=current_row, column=1).font = Font(name="Calibri", bold=True, size=12, color="C00000")
    current_row += 1
    ws.cell(row=current_row, column=1, value="(Add your revenue growth, margin, and other assumptions below)")
    ws.cell(row=current_row, column=1).font = Font(name="Calibri", italic=True, size=10, color="808080")

    placeholders = [
        "Revenue Growth %",
        "Gross Margin %",
        "Operating Margin %",
        "Tax Rate %",
        "CapEx % of Revenue",
        "D&A % of Revenue",
    ]
    current_row += 1
    for label in placeholders:
        ws.cell(row=current_row, column=1, value=label).font = DATA_FONT
        current_row += 1

    ws.column_dimensions["A"].width = 45
    for col_idx in range(2, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def export_to_excel(statements, company_name, ticker, output_dir=None):
    """Export financial statements to a formatted Excel workbook.

    Args:
        statements: Dict mapping statement names to DataFrames.
        company_name: Company name for titles.
        ticker: Ticker symbol (used in filename).
        output_dir: Directory to save the file. Defaults to config OUTPUT_DIR.

    Returns:
        Path to the saved Excel file.
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Add summary first
    _add_summary_sheet(wb, statements, company_name)

    # Add each financial statement as a tab
    for sheet_name, df in statements.items():
        ws = wb.create_sheet(sheet_name)
        if df is not None and not df.empty:
            _style_sheet(ws, df, f"{company_name} — {sheet_name}")
        else:
            ws.cell(row=1, column=1, value=f"No data available for {sheet_name}")

    filepath = os.path.join(output_dir, f"{ticker.upper()}_3_statement_model.xlsx")
    wb.save(filepath)
    return filepath
